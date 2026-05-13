"""Multi-dimensional DFlash speculation analysis.

Builds a structured Excel workbook with:

  * README          analytical formulas + interpretive notes
  * main_effects    six 1D sweeps (each parameter held against speedup,
                    others fixed at defaults). Answers 'what's the marginal
                    effect of axis X?'
  * pair_*          three 2D heatmaps for the most informative interactions
                    (ep × batch · ep × accept · past × batch)
  * pro_vs_flash    side-by-side speedup matrix for the two main models
  * canonical_vs_hybrid  same for the two draft variants
  * comm_sensitivity  ep × inter_rank_bw × overlap grid showing where comm
                      becomes the bottleneck
  * deepep_impact   focused (ep × batch) comparison of comm-off /
                    NCCL-serial (overlap=0) / DeepEP-grade (overlap=0.7).
                    Quantifies the DeepEP value.
  * hw_compare      same workload on H100 vs Ascend950PR
  * fp4_experts     bf16 vs fp4 expert weight dtype on V4-Pro (the published
                    config uses fp4)
  * rounds          flat record dump for pivot-table work in Excel
"""
from __future__ import annotations

import argparse
import math
import sys
import time
from dataclasses import dataclass
from itertools import product
from pathlib import Path

import torch

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from alloy.examples.roofline.dsv4_pro import build_v4_pro_config
from alloy.examples.roofline.dsv4_flash import build_v4_flash_config
from alloy.examples.roofline.dflash.dsv4_pro_dflash import (
    build_dsv4_pro_dflash_config,
)
from alloy.examples.roofline.dflash.dsv4_pro_dflash_hybrid import (
    build_dsv4_pro_dflash_hybrid_config,
)
from alloy.examples.roofline.dflash.dsv4_flash_dflash import (
    build_dsv4_flash_dflash_config,
)
from alloy.examples.roofline.dflash.dsv4_flash_dflash_hybrid import (
    build_dsv4_flash_dflash_hybrid_config,
)
from alloy.roofline import (
    CustomHardware,
    get_hardware,
    roofline_decode,
    roofline_mini_prefill,
    roofline_dflash_draft_forward,
)

DTYPE_BYTES = {"bf16": 2, "fp16": 2, "fp8": 1, "fp4": 0.5, "fp32": 4}


# --------------------------------------------------------------------------- #
# Pair / hardware registry
# --------------------------------------------------------------------------- #

MAIN_BUILDERS = {
    "DSV4-Pro":   build_v4_pro_config,
    "DSV4-Flash": build_v4_flash_config,
}

DRAFT_BUILDERS = {
    ("DSV4-Pro",   "canonical"): build_dsv4_pro_dflash_config,
    ("DSV4-Pro",   "hybrid"):    build_dsv4_pro_dflash_hybrid_config,
    ("DSV4-Flash", "canonical"): build_dsv4_flash_dflash_config,
    ("DSV4-Flash", "hybrid"):    build_dsv4_flash_dflash_hybrid_config,
}


def get_hw(name):
    if name == "my-device":
        return CustomHardware(name="my-device", hbm_bandwidth=8e12,
                              fp16=1000e12, bf16=1000e12, fp8=2000e12)
    return get_hardware(name)


# --------------------------------------------------------------------------- #
# Communication cost
# --------------------------------------------------------------------------- #

def comm_time(target_cfg, n_tokens, dtype, ep, inter_bw_gbs, overlap, enabled):
    """Per speculation round comm time (seconds).

    Formula (per MoE layer, dispatch + combine):
        2 * N * K * (ep-1)/ep * H * es  bytes off-rank
    Total over all MoE layers in the target, then divided by BW, then
    scaled by (1 - overlap).
    """
    if not enabled or ep <= 1:
        return 0.0
    es = DTYPE_BYTES[dtype]
    K = target_cfg.num_experts_per_tok
    H = target_cfg.hidden_size
    n_moe = sum(1 for f in target_cfg.ffn_types
                if f in ("dsv4_moe", "dsv4_hash_moe"))
    off_rank = (ep - 1) / ep
    per_layer_bytes = 2 * n_tokens * K * off_rank * H * es
    total_bytes = n_moe * per_layer_bytes
    t = total_bytes / (inter_bw_gbs * 1e9)
    return t * (1.0 - overlap)


# --------------------------------------------------------------------------- #
# One spec round → SpecRecord
# --------------------------------------------------------------------------- #

@dataclass
class SpecRecord:
    main: str
    draft: str
    hw: str
    batch: int
    past: int
    block: int
    accept: float
    ep: int
    routing: float
    comm_on: bool
    comm_bw: float
    overlap: float
    dtype: str
    expert_dtype: str

    # Derived
    t_ar: float
    t_verify: float
    t_draft: float
    t_comm: float
    t_round: float
    ar_tps: float
    dflash_tps: float
    speedup: float


# Cache of (main, draft) config objects so we don't rebuild every call.
_CONFIG_CACHE: dict[tuple, tuple] = {}


def _configs_for(main: str, draft: str):
    key = (main, draft)
    if key not in _CONFIG_CACHE:
        target = MAIN_BUILDERS[main]()
        d_builder = DRAFT_BUILDERS[(main, draft)]
        _CONFIG_CACHE[key] = (target, d_builder())
    return _CONFIG_CACHE[key]


def evaluate(
    *, main: str, draft: str, hw: str,
    batch: int, past: int, block: int, accept: float,
    ep: int = 1, routing: float = 1.0,
    comm_on: bool = True, comm_bw: float = 900.0, overlap: float = 0.0,
    dtype: str = "bf16", expert_dtype: str = "bf16",
) -> SpecRecord:
    hw_name = hw
    target_cfg, draft_cfg = _configs_for(main, draft)
    hw = get_hw(hw_name)
    dt = {"bf16": torch.bfloat16, "fp16": torch.float16, "fp8": torch.float8_e4m3fn}[dtype]

    # NB: alloy's roofline doesn't expose routing_collision / ep / expert_dtype
    # as kwargs; the HTML/JS does. For the sweep tool we patch the target
    # config temporarily with scaled n_routed_experts to approximate ep+routing
    # effects on MoE expert byte traffic. This keeps the analysis directionally
    # correct without forking the spec code.
    saved = {}
    if hasattr(target_cfg, "n_routed_experts"):
        # Effective unique experts loaded per rank (ANY-routing upper bound)
        E = target_cfg.n_routed_experts
        K = target_cfg.num_experts_per_tok
        # AR (Q=1): nTokens*K = batch*K; verify uses batch*block*K
        # We can't easily 'monkeypatch' per-call. Instead, scale n_routed_experts
        # by min(1, routing) * (1/ep) AND clamp to keep semantics consistent.
        # Simpler and exactly matches our HTML formula: replace E with a
        # scaled E' such that downstream min(E', N*K) gives the per-rank
        # unique count we want. We do this only on a *copy* of the config.
        pass

    ar = roofline_decode(target_cfg, batch=batch, kv_cache_len=past,
                          dtype=dt, hardware=hw)
    verify = roofline_mini_prefill(target_cfg, batch=batch, chunk_len=block,
                                    kv_cache_len=past, dtype=dt, hardware=hw)
    ctx_len = max(1, int(round(accept + 1)))
    draft_r = roofline_dflash_draft_forward(
        draft_cfg, batch=batch, block_size=block, ctx_len=ctx_len,
        past_cache_len=past, hardware=hw, dtype=dt,
    )

    # MoE-expert-byte adjustment for routing + EP (post-hoc):
    # alloy used unique_E = min(E, n_tokens * K). We replace that with
    # unique_per_rank = max(1, ceil(unique_E * routing / ep)). The replacement
    # changes only the routed_w portion of the byte total; we estimate the
    # delta and adjust verify/AR bytes accordingly.
    def adjust(report, n_tokens):
        nonlocal target_cfg
        if not hasattr(target_cfg, "n_routed_experts"):
            return report.roofline_time_s, report.bottleneck
        # Identify expert byte contribution: per-expert weight bytes
        from alloy.roofline.specs_ffn import _swiglu_weight_bytes
        from alloy.roofline.specs import dtype_size
        H = target_cfg.hidden_size
        I = target_cfg.intermediate_size
        bias = bool(getattr(target_cfg, "mlp_bias", False))
        expert_es = DTYPE_BYTES[expert_dtype]
        per_expert_w = _swiglu_weight_bytes(H, I, expert_es, bias)
        # alloy assumed unique = min(E, nTokens * K). bf16-quantization shift
        # we apply by scaling the routed bytes contribution.
        alloy_unique = min(target_cfg.n_routed_experts, n_tokens * target_cfg.num_experts_per_tok)
        # Effective unique per rank with routing collision + EP
        our_unique = max(1, math.ceil(alloy_unique * routing / ep))
        # alloy used dtype bytes for expert weights; we override with expert_dtype
        alloy_per_expert = _swiglu_weight_bytes(H, I, dtype_size(dt), bias)
        # Per-layer delta in bytes (across all MoE layers)
        n_moe = sum(1 for f in target_cfg.ffn_types
                    if f in ("dsv4_moe", "dsv4_hash_moe"))
        old_routed = n_moe * alloy_unique * alloy_per_expert
        new_routed = n_moe * our_unique * per_expert_w
        # Also adjust shared expert bytes for expert_dtype change
        old_shared = n_moe * alloy_per_expert
        new_shared = n_moe * per_expert_w
        adjusted_bytes = report.total_bytes - old_routed - old_shared + new_routed + new_shared
        peak = hw.peak_flops[dt]
        new_mem_time = adjusted_bytes / hw.hbm_bandwidth
        old_compute_time = report.total_flops / peak
        new_wall = max(old_compute_time, new_mem_time)
        return new_wall, ("compute" if old_compute_time > new_mem_time else "memory")

    t_ar, _ = adjust(ar, batch * 1)
    t_verify, _ = adjust(verify, batch * block)
    t_draft = draft_r.roofline_time_s  # draft has no MoE → no adjustment

    n_tokens_verify = batch * block
    t_comm = comm_time(target_cfg, n_tokens_verify, dtype,
                       ep, comm_bw, overlap, comm_on)

    t_round = t_draft + t_verify + t_comm
    tokens_per_round = accept + 1.0
    dflash_tps = batch * tokens_per_round / t_round
    ar_tps = batch * 1.0 / t_ar
    speedup = dflash_tps / ar_tps if ar_tps > 0 else 0.0

    return SpecRecord(
        main=main, draft=draft, hw=hw_name,
        batch=batch, past=past, block=block, accept=accept,
        ep=ep, routing=routing, comm_on=comm_on, comm_bw=comm_bw,
        overlap=overlap, dtype=dtype, expert_dtype=expert_dtype,
        t_ar=t_ar, t_verify=t_verify, t_draft=t_draft, t_comm=t_comm,
        t_round=t_round, ar_tps=ar_tps, dflash_tps=dflash_tps,
        speedup=speedup,
    )


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #

def sweep_grid(axes_dict, defaults):
    """Yield records over a grid where each axis in axes_dict varies, others
    pinned to defaults."""
    for combo in product(*axes_dict.values()):
        params = {**defaults, **dict(zip(axes_dict.keys(), combo))}
        yield evaluate(**params)


# --------------------------------------------------------------------------- #
# Excel writer
# --------------------------------------------------------------------------- #

def write_xlsx(path, sheet_builders, defaults_text):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    WIN_HIGH = PatternFill("solid", fgColor="86efac")     # speedup >= 2
    WIN = PatternFill("solid", fgColor="d4f4d4")          # 1.5-2
    EQUAL = PatternFill("solid", fgColor="fff7cc")        # 1-1.5
    LOSE_MILD = PatternFill("solid", fgColor="ffe4d4")    # 0.7-1
    LOSE = PatternFill("solid", fgColor="fce4e4")         # <0.7
    HDR = PatternFill("solid", fgColor="d9e1f2")
    HDR_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN = Side(style="thin", color="cccccc")
    BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def color_for(speedup):
        if speedup >= 2.0: return WIN_HIGH
        if speedup >= 1.5: return WIN
        if speedup >= 1.0: return EQUAL
        if speedup >= 0.7: return LOSE_MILD
        return LOSE

    wb = Workbook()
    first = wb.active
    first.title = "README"
    write_readme(first, defaults_text, HDR, HDR_FONT, LEFT)

    for name, builder in sheet_builders.items():
        ws = wb.create_sheet(name[:31])
        builder(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR)

    wb.save(path)
    print(f"Excel written: {path}")


def write_readme(ws, defaults_text, HDR, HDR_FONT, LEFT):
    from openpyxl.styles import Font
    lines = [
        ("DFlash Multi-Dimensional Analysis — README", True, 14),
        ("", False, 11),
        (f"Generated by tools/dflash/multi_dim_analysis.py", False, 10),
        ("", False, 11),
        ("=== Speedup formula ===", True, 12),
        ("speedup = (accept_len + 1) × t_AR / (t_draft + t_verify + t_comm)", False, 11),
        ("", False, 11),
        ("    t_AR     = AR target decode at given past_cache (Q=1)", False, 11),
        ("    t_verify = target mini-prefill on the block (Q=block_size)", False, 11),
        ("    t_draft  = DFlash draft forward over block_size with ctx_len ≈ accept+1", False, 11),
        ("    t_comm   = target-side all-to-all per round, scaled by (1 - overlap_factor)", False, 11),
        ("", False, 11),
        ("=== Break-even accept length ===", True, 12),
        ("accept_BE = (t_draft + t_verify + t_comm) / t_AR  -  1", False, 11),
        ("Spec wins iff actual avg_accept > accept_BE.", False, 11),
        ("", False, 11),
        ("=== MoE-dominated approximation ===", True, 12),
        ("When expert weights dominate t_AR and t_verify:", False, 11),
        ("    t_verify / t_AR ≈ min(E, B·block·K·collision) / min(E, K·collision)  [single rank]", False, 11),
        ("Per rank with EP=R:", False, 11),
        ("    unique_AR_per_rank     = max(1, min(E, K·collision) / R)", False, 11),
        ("    unique_verify_per_rank = max(1, min(E, B·block·K·collision) / R)", False, 11),
        ("So the ratio depends on EP only via the clamp-to-1 cutoff.", False, 11),
        ("", False, 11),
        ("=== EP 'win' threshold ===", True, 12),
        ("If you want expert byte traffic to be equal on AR and verify (so no", False, 11),
        ("expert-load handicap):", False, 11),
        ("    R ≥ B · block · K · collision", False, 11),
        ("e.g. B=1, block=16, K=6, collision=1.0 → R ≥ 96", False, 11),
        ("e.g. B=1, block=16, K=6, collision=0.3 → R ≥ 29  (smaller EP suffices)", False, 11),
        ("", False, 11),
        ("=== Communication cost ===", True, 12),
        ("Per MoE layer, dispatch + combine all-to-all:", False, 11),
        ("    bytes = 2 · N · K · (EP-1)/EP · H · es", False, 11),
        ("Total per round = num_MoE_layers · per_layer · (1 - overlap_factor) / inter_rank_BW", False, 11),
        ("", False, 11),
        ("=== DeepEP impact ===", True, 12),
        ("DeepEP overlaps dispatch / expert-compute / combine within a single MoE layer.", False, 11),
        ("Published numbers achieve ~70% overlap on H100.", False, 11),
        ("Modelled as: t_comm_effective = t_comm_raw · (1 - overlap)", False, 11),
        ("  overlap=0.0  : NCCL serial (vLLM / SGLang default)", False, 11),
        ("  overlap=0.7  : DeepEP-grade", False, 11),
        ("  overlap=1.0  : perfect (upper bound, not realisable)", False, 11),
        ("", False, 11),
        ("=== Cell colour legend (speedup) ===", True, 12),
        ("  ≥ 2.0x   dark green", False, 11),
        ("  1.5–2.0  light green", False, 11),
        ("  1.0–1.5  yellow", False, 11),
        ("  0.7–1.0  light red", False, 11),
        ("  < 0.7    red", False, 11),
        ("", False, 11),
        ("=== Caveats ===", True, 12),
        ("(1) Roofline = peak HBM + peak FLOPS, no kernel overhead / scheduling.", False, 11),
        ("(2) Hybrid draft modelling is approximate — it uses DFlashAttentionSpec", False, 11),
        ("    (plain Qwen3-dim, non-causal + ctx_len cat) for all draft layers, NOT", False, 11),
        ("    a true MLA + CSA/HCA + DFlash spec. The MLA savings aren't captured.", False, 11),
        ("(3) MoE expert bytes use worst-case 'any-routing' upper bound", False, 11),
        ("    min(E, N·K) — realistic routing locality typically gives 0.2-0.4 collision.", False, 11),
        ("(4) Communication is modelled as additive serial cost (modulo overlap).", False, 11),
        ("    Real all-to-all has start-up latency that this doesn't capture.", False, 11),
        ("", False, 11),
        ("=== Default workload axis values ===", True, 12),
        (defaults_text, False, 11),
    ]
    for r, (text, bold, size) in enumerate(lines, 1):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = LEFT
    ws.column_dimensions["A"].width = 110


# --------------------------------------------------------------------------- #
# Sheet 2: Main effects (1D sweeps)
# --------------------------------------------------------------------------- #

def make_main_effects_sheet(defaults):
    axes = {
        "batch":   [1, 4, 16, 64, 256],
        "past":    [8192, 32768, 131072, 524288, 1048576, 4194304],
        "block":   [4, 8, 16, 32, 64],
        "accept":  [0.5, 1.0, 1.44, 2.0, 2.44, 3.0, 4.0],
        "ep":      [1, 2, 4, 8, 16, 32, 64, 128],
        "routing": [0.1, 0.2, 0.3, 0.5, 0.7, 1.0],
    }

    def fmt_axis(v, axis):
        if axis == "past": return _scale_seq(v)
        if isinstance(v, float): return f"{v:.2f}"
        return str(v)

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value="Main effects — speedup of varying ONE axis while others stay at defaults").font = Font(bold=True, size=12)
        row = 3
        for axis, values in axes.items():
            ws.cell(row=row, column=1, value=f"axis: {axis}").font = HDR_FONT
            row += 1
            # Header
            ws.cell(row=row, column=1, value="value").font = HDR_FONT
            ws.cell(row=row, column=1).fill = HDR
            for j, h in enumerate(["t_AR (ms)", "t_round (ms)", "AR tps", "DFlash tps", "speedup"], 2):
                c = ws.cell(row=row, column=j, value=h); c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
            row += 1
            for v in values:
                rec = evaluate(**{**defaults, axis: v})
                ws.cell(row=row, column=1, value=fmt_axis(v, axis))
                ws.cell(row=row, column=2, value=round(rec.t_ar * 1000, 3))
                ws.cell(row=row, column=3, value=round(rec.t_round * 1000, 3))
                ws.cell(row=row, column=4, value=round(rec.ar_tps, 2))
                ws.cell(row=row, column=5, value=round(rec.dflash_tps, 2))
                c = ws.cell(row=row, column=6, value=round(rec.speedup, 3))
                c.fill = color_for(rec.speedup)
                for j in range(1, 7):
                    ws.cell(row=row, column=j).border = BDR
                row += 1
            row += 1  # blank between axes
        # widths
        for j, w in enumerate([14, 14, 14, 14, 14, 12], 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(j)].width = w
    return build


def _scale_seq(s: int) -> str:
    if s % (1024 * 1024) == 0: return f"{s // (1024 * 1024)}M"
    if s % 1024 == 0: return f"{s // 1024}K"
    return str(s)


# --------------------------------------------------------------------------- #
# Pair heatmap helper
# --------------------------------------------------------------------------- #

def make_pair_heatmap(defaults, row_axis, row_vals, col_axis, col_vals,
                     metric_fn, metric_label, title):
    def fmt(v, axis):
        if axis == "past": return _scale_seq(v)
        if isinstance(v, float): return f"{v:.2f}"
        return str(v)

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value=f"defaults: " + _summarise_defaults(defaults))
        ws.cell(row=3, column=1, value=f"cell = {metric_label}")
        header_row = 5
        ws.cell(row=header_row, column=1, value=f"{row_axis} ↓  /  {col_axis} →").font = HDR_FONT
        ws.cell(row=header_row, column=1).fill = HDR
        for j, cv in enumerate(col_vals, 2):
            c = ws.cell(row=header_row, column=j, value=fmt(cv, col_axis))
            c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
        row_idx = header_row + 1
        for rv in row_vals:
            c = ws.cell(row=row_idx, column=1, value=fmt(rv, row_axis))
            c.font = HDR_FONT; c.fill = HDR
            for j, cv in enumerate(col_vals, 2):
                rec = evaluate(**{**defaults, row_axis: rv, col_axis: cv})
                val = metric_fn(rec)
                cell = ws.cell(row=row_idx, column=j, value=val)
                if metric_label.startswith("speedup"):
                    cell.fill = color_for(val)
                cell.alignment = CENTER
                cell.border = BDR
            row_idx += 1
        # widths
        ws.column_dimensions["A"].width = 18
        from openpyxl.utils import get_column_letter
        for j in range(2, 2 + len(col_vals)):
            ws.column_dimensions[get_column_letter(j)].width = 12
    return build


def _summarise_defaults(d):
    return (
        f"main={d['main']}  draft={d['draft']}  hw={d['hw']}  "
        f"B={d['batch']}  P={_scale_seq(d['past'])}  block={d['block']}  "
        f"accept={d['accept']}  ep={d['ep']}  routing={d['routing']}  "
        f"comm_on={d['comm_on']}  bw={d['comm_bw']}  overlap={d['overlap']}  "
        f"dtype={d['dtype']}  expert_dtype={d['expert_dtype']}"
    )


# --------------------------------------------------------------------------- #
# Sheet 6: Pro vs Flash
# --------------------------------------------------------------------------- #

def make_pro_vs_flash_sheet(defaults):
    batches = [1, 8, 64]
    accepts = [0.7, 1.44, 2.44]

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value="Main model comparison — DSV4-Pro vs DSV4-Flash speedup").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value="defaults: " + _summarise_defaults(defaults))
        ws.cell(row=3, column=1, value="cell = speedup (DFlash tps / AR tps)")
        row = 5
        for main in ("DSV4-Pro", "DSV4-Flash"):
            ws.cell(row=row, column=1, value=main).font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="B \\ accept").font = HDR_FONT
            ws.cell(row=row, column=1).fill = HDR
            for j, a in enumerate(accepts, 2):
                c = ws.cell(row=row, column=j, value=f"acc={a:.2f}")
                c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
            row += 1
            for B in batches:
                ws.cell(row=row, column=1, value=B).fill = HDR
                for j, a in enumerate(accepts, 2):
                    rec = evaluate(**{**defaults, "main": main, "batch": B, "accept": a})
                    c = ws.cell(row=row, column=j, value=round(rec.speedup, 3))
                    c.fill = color_for(rec.speedup); c.alignment = CENTER; c.border = BDR
                row += 1
            row += 1
        ws.column_dimensions["A"].width = 16
    return build


# --------------------------------------------------------------------------- #
# Sheet 7: canonical vs hybrid
# --------------------------------------------------------------------------- #

def make_canonical_vs_hybrid_sheet(defaults):
    pasts = [8192, 131072, 1048576, 4194304]
    batches = [1, 8, 64]

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value="Draft variant comparison — canonical vs hybrid speedup").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value="defaults: " + _summarise_defaults(defaults))
        ws.cell(row=3, column=1, value="cell = speedup;  ⚠ hybrid uses approximate spec (no MLA savings captured)")
        row = 5
        for draft in ("canonical", "hybrid"):
            ws.cell(row=row, column=1, value=f"draft = {draft}").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="B \\ past").font = HDR_FONT
            ws.cell(row=row, column=1).fill = HDR
            for j, p in enumerate(pasts, 2):
                c = ws.cell(row=row, column=j, value=_scale_seq(p))
                c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
            row += 1
            for B in batches:
                ws.cell(row=row, column=1, value=B).fill = HDR
                for j, p in enumerate(pasts, 2):
                    rec = evaluate(**{**defaults, "draft": draft, "batch": B, "past": p})
                    c = ws.cell(row=row, column=j, value=round(rec.speedup, 3))
                    c.fill = color_for(rec.speedup); c.alignment = CENTER; c.border = BDR
                row += 1
            row += 1
        ws.column_dimensions["A"].width = 16
    return build


# --------------------------------------------------------------------------- #
# Sheet 8: Comm sensitivity
# --------------------------------------------------------------------------- #

def make_comm_sensitivity_sheet(defaults):
    eps = [1, 4, 16, 64, 256]
    bws = [12.5, 25, 50, 100, 200, 900]  # GbE → NVLink span

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value="Communication sensitivity — speedup over (EP × inter_rank_bw) at overlap=0").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value="defaults: " + _summarise_defaults(defaults))
        ws.cell(row=3, column=1, value="cell = speedup at the (EP, bw) combo")
        header_row = 5
        ws.cell(row=header_row, column=1, value="EP \\ bw (GB/s)").font = HDR_FONT
        ws.cell(row=header_row, column=1).fill = HDR
        for j, bw in enumerate(bws, 2):
            c = ws.cell(row=header_row, column=j, value=bw)
            c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
        r = header_row + 1
        for ep in eps:
            ws.cell(row=r, column=1, value=ep).fill = HDR
            for j, bw in enumerate(bws, 2):
                rec = evaluate(**{**defaults, "ep": ep, "comm_bw": bw, "overlap": 0.0})
                c = ws.cell(row=r, column=j, value=round(rec.speedup, 3))
                c.fill = color_for(rec.speedup); c.alignment = CENTER; c.border = BDR
            r += 1
        ws.column_dimensions["A"].width = 16
    return build


# --------------------------------------------------------------------------- #
# Sheet 9: DeepEP impact
# --------------------------------------------------------------------------- #

def make_deepep_impact_sheet(defaults):
    eps = [1, 4, 16, 64, 128, 256]
    batches = [1, 4, 16, 64, 256]

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value="DeepEP impact — comm-off vs NCCL serial (overlap=0) vs DeepEP-grade (overlap=0.7)").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value="defaults: " + _summarise_defaults(defaults))
        ws.cell(row=3, column=1, value="Three side-by-side blocks; cell = speedup")
        row = 5
        scenarios = [
            ("(A) comm OFF (treat comm as free)", {"comm_on": False, "overlap": 0.0}),
            ("(B) NCCL serial (overlap=0)",        {"comm_on": True,  "overlap": 0.0}),
            ("(C) DeepEP (overlap=0.7)",            {"comm_on": True,  "overlap": 0.7}),
        ]
        for title, overrides in scenarios:
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="EP \\ B").font = HDR_FONT
            ws.cell(row=row, column=1).fill = HDR
            for j, B in enumerate(batches, 2):
                c = ws.cell(row=row, column=j, value=f"B={B}")
                c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
            row += 1
            for ep in eps:
                ws.cell(row=row, column=1, value=ep).fill = HDR
                for j, B in enumerate(batches, 2):
                    rec = evaluate(**{**defaults, **overrides, "ep": ep, "batch": B})
                    c = ws.cell(row=row, column=j, value=round(rec.speedup, 3))
                    c.fill = color_for(rec.speedup); c.alignment = CENTER; c.border = BDR
                row += 1
            row += 1
        # NCCL → DeepEP delta block
        ws.cell(row=row, column=1, value="(D) DeepEP delta = (C - B) — pure DeepEP value-add").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="EP \\ B").font = HDR_FONT
        ws.cell(row=row, column=1).fill = HDR
        for j, B in enumerate(batches, 2):
            c = ws.cell(row=row, column=j, value=f"B={B}")
            c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
        row += 1
        for ep in eps:
            ws.cell(row=row, column=1, value=ep).fill = HDR
            for j, B in enumerate(batches, 2):
                rb = evaluate(**{**defaults, "comm_on": True, "overlap": 0.0, "ep": ep, "batch": B})
                rc = evaluate(**{**defaults, "comm_on": True, "overlap": 0.7, "ep": ep, "batch": B})
                delta = rc.speedup - rb.speedup
                c = ws.cell(row=row, column=j, value=round(delta, 3))
                # Colour: orange shades by magnitude
                if delta >= 0.5:    fill_hex = "fdba74"  # strong DeepEP win
                elif delta >= 0.1:  fill_hex = "fed7aa"  # mild win
                elif delta >= 0.01: fill_hex = "fef3c7"  # marginal
                else:               fill_hex = "f1f5f9"  # negligible
                from openpyxl.styles import PatternFill as PF
                c.fill = PF("solid", fgColor=fill_hex)
                c.alignment = CENTER; c.border = BDR
            row += 1
        ws.column_dimensions["A"].width = 14
    return build


# --------------------------------------------------------------------------- #
# Sheet 10: HW compare
# --------------------------------------------------------------------------- #

def make_hw_compare_sheet(defaults):
    hardwares = ["H100", "Ascend950PR", "my-device"]
    eps = [1, 8, 64]
    pasts = [8192, 131072, 1048576]

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value="Hardware comparison — same workload on H100 / 950PR / my-device").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value="defaults: " + _summarise_defaults(defaults))
        row = 4
        for hw_name in hardwares:
            ws.cell(row=row, column=1, value=hw_name).font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="EP \\ past").font = HDR_FONT
            ws.cell(row=row, column=1).fill = HDR
            for j, p in enumerate(pasts, 2):
                c = ws.cell(row=row, column=j, value=_scale_seq(p))
                c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
            row += 1
            for ep in eps:
                ws.cell(row=row, column=1, value=ep).fill = HDR
                for j, p in enumerate(pasts, 2):
                    rec = evaluate(**{**defaults, "hw": hw_name, "ep": ep, "past": p})
                    c = ws.cell(row=row, column=j, value=round(rec.speedup, 3))
                    c.fill = color_for(rec.speedup); c.alignment = CENTER; c.border = BDR
                row += 1
            row += 1
        ws.column_dimensions["A"].width = 16
    return build


# --------------------------------------------------------------------------- #
# Sheet 11: FP4 experts
# --------------------------------------------------------------------------- #

def make_fp4_experts_sheet(defaults):
    eps = [1, 4, 16, 64, 256]
    batches = [1, 8, 64]

    def build(ws, color_for, HDR, HDR_FONT, CENTER, LEFT, BDR):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1, value="Expert dtype comparison — BF16 vs FP4 experts (V4-Pro uses FP4)").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value="defaults: " + _summarise_defaults(defaults))
        row = 4
        for expert_dt in ("bf16", "fp4"):
            ws.cell(row=row, column=1, value=f"expert_dtype = {expert_dt}").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="EP \\ B").font = HDR_FONT
            ws.cell(row=row, column=1).fill = HDR
            for j, B in enumerate(batches, 2):
                c = ws.cell(row=row, column=j, value=f"B={B}")
                c.font = HDR_FONT; c.fill = HDR; c.alignment = CENTER
            row += 1
            for ep in eps:
                ws.cell(row=row, column=1, value=ep).fill = HDR
                for j, B in enumerate(batches, 2):
                    rec = evaluate(**{**defaults, "expert_dtype": expert_dt, "ep": ep, "batch": B})
                    c = ws.cell(row=row, column=j, value=round(rec.speedup, 3))
                    c.fill = color_for(rec.speedup); c.alignment = CENTER; c.border = BDR
                row += 1
            row += 1
        ws.column_dimensions["A"].width = 16
    return build


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", default="D:/work/model_gym/dflash_multi_dim.xlsx",
                        help="Output path. Default: D:/work/model_gym/dflash_multi_dim.xlsx")
    parser.add_argument("--main", default="DSV4-Pro", choices=list(MAIN_BUILDERS))
    parser.add_argument("--draft", default="hybrid", choices=["canonical", "hybrid"])
    parser.add_argument("--hw", default="H100")
    parser.add_argument("--batch", type=int, default=1)
    parser.add_argument("--past", type=int, default=131072)
    parser.add_argument("--block", type=int, default=16)
    parser.add_argument("--accept", type=float, default=1.44)
    parser.add_argument("--ep", type=int, default=1)
    parser.add_argument("--routing", type=float, default=1.0)
    parser.add_argument("--no-comm", action="store_true", help="Default sweep with comm off")
    parser.add_argument("--bw", type=float, default=900.0)
    parser.add_argument("--overlap", type=float, default=0.0)
    parser.add_argument("--dtype", default="bf16")
    parser.add_argument("--expert-dtype", default="bf16")
    args = parser.parse_args()

    defaults = dict(
        main=args.main, draft=args.draft, hw=args.hw,
        batch=args.batch, past=args.past, block=args.block, accept=args.accept,
        ep=args.ep, routing=args.routing,
        comm_on=not args.no_comm, comm_bw=args.bw, overlap=args.overlap,
        dtype=args.dtype, expert_dtype=args.expert_dtype,
    )
    defaults_text = _summarise_defaults(defaults)
    print(f"Defaults: {defaults_text}")
    t0 = time.perf_counter()

    sheets = {
        "main_effects":         make_main_effects_sheet(defaults),
        "pair_ep_batch":        make_pair_heatmap(
            defaults, "ep", [1, 2, 4, 8, 16, 32, 64, 128, 256],
            "batch", [1, 4, 16, 64, 256],
            lambda r: round(r.speedup, 3), "speedup",
            "Pair: EP × batch — speedup (others at defaults)"),
        "pair_ep_accept":       make_pair_heatmap(
            defaults, "ep", [1, 4, 16, 64, 256],
            "accept", [0.5, 1.0, 1.44, 2.0, 2.44, 3.0, 4.0],
            lambda r: round(r.speedup, 3), "speedup",
            "Pair: EP × accept — speedup"),
        "pair_past_batch":      make_pair_heatmap(
            defaults, "past", [8192, 32768, 131072, 524288, 1048576, 4194304],
            "batch", [1, 4, 16, 64, 256],
            lambda r: round(r.speedup, 3), "speedup",
            "Pair: past_cache × batch — speedup"),
        "pair_routing_ep":      make_pair_heatmap(
            defaults, "routing", [0.1, 0.2, 0.3, 0.5, 0.7, 1.0],
            "ep", [1, 4, 16, 64, 256],
            lambda r: round(r.speedup, 3), "speedup",
            "Pair: routing_collision × EP — speedup"),
        "pro_vs_flash":         make_pro_vs_flash_sheet(defaults),
        "canonical_vs_hybrid":  make_canonical_vs_hybrid_sheet(defaults),
        "comm_sensitivity":     make_comm_sensitivity_sheet(defaults),
        "deepep_impact":        make_deepep_impact_sheet(defaults),
        "hw_compare":           make_hw_compare_sheet(defaults),
        "fp4_experts":          make_fp4_experts_sheet(defaults),
    }
    write_xlsx(args.xlsx, sheets, defaults_text)
    elapsed = time.perf_counter() - t0
    print(f"Done in {elapsed:.1f}s")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
