"""Cross-model workload sweep for DSV4 + DFlash drafts.

Iterates over a 6 models * N hardware * 4 batch sizes * 5 sequence
lengths * 3 modes grid and prints / dumps the resulting roofline stats.
Produces:

  * Per-(hardware, dtype) ridge point header.
  * Per-(hardware, mode) **summary** matrices: rows = (model, batch),
    cols = seq_len. Each cell shows ``time / tok-per-s / bound`` for
    the whole model.
  * Optional per-(hardware, model, batch, seq_len, mode) **component**
    breakdown via ``--components`` — each block lists every
    embedding / mhc / mixer / ffn / norm / lm_head ModuleStat with its
    FLOPs, bytes, AI, bound, and time.
  * Optional CSV dump (``--csv path``) with one row per
    (hardware, model, batch, seq_len, mode) plus a separate
    ``<csv>.components.csv`` with one row per ModuleStat.

Conventions:

  * ``prefill``: Q = seq_len, P = 0 (cold start, full prompt at once).
  * ``mini-prefill``: Q = ``--chunk-len`` (default 512), P = seq_len
    (one chunk of chunked prefill against an existing context).
  * ``decode``: Q = 1, P = seq_len (generate one token with a seq_len
    cache).

Theoretical per-component ridge point: as the user noted, this depends
only on the hardware (``peak_flops / hbm_bandwidth``) and dtype, NOT on
the operator. So we print it once per (hardware, dtype) at the top of
each hardware section. Components are still tagged with their per-row
bound ('C' compute / 'M' memory) for at-a-glance reading.

Usage::

    # default — summary on H100 + Ascend950PR + my-device equivalent placeholder
    python -m alloy.tools.dflash.analyze_workloads

    # include per-component breakdowns
    python -m alloy.tools.dflash.analyze_workloads --components

    # narrow the sweep
    python -m alloy.tools.dflash.analyze_workloads \
        --models DSV4-Pro,DSV4-Pro-DFlash \
        --batches 1,8 --seq-lens 8192,131072 \
        --hardware H100

    # dump full data for downstream analysis
    python -m alloy.tools.dflash.analyze_workloads --csv /tmp/dflash_sweep.csv

    # human-readable Excel workbook (one sheet per (hw, mode), colour-coded
    # by bound, plus a flat 'totals' sheet with auto-filter)
    python -m alloy.tools.dflash.analyze_workloads --xlsx /tmp/dflash_sweep.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Callable

import torch

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from alloy.roofline import (
    CustomHardware,
    RooflineReport,
    get_hardware,
    roofline_decode,
    roofline_mini_prefill,
    roofline_prefill,
)

# Single source of truth for each model's config — reuse the example builders.
from alloy.examples.roofline.dsv4_pro import build_v4_pro_config
from alloy.examples.roofline.dsv4_flash import build_v4_flash_config
from alloy.examples.roofline.dflash.dsv4_pro_dflash import (
    build_dsv4_pro_dflash_config,
)
from alloy.examples.roofline.dflash.dsv4_pro_dflash_hybrid import (
    build_dsv4_pro_dflash_hybrid_config,
)
from alloy.examples.roofline.dflash.dsv4_flash_dflash import (
    build_dsv4_flash_dflash_config,
)
from alloy.examples.roofline.dflash.dsv4_flash_dflash_hybrid import (
    build_dsv4_flash_dflash_hybrid_config,
)


# --------------------------------------------------------------------------- #
# Sweep axes
# --------------------------------------------------------------------------- #

ModelBuilder = Callable[[], object]

ALL_MODELS: "OrderedDict[str, ModelBuilder]" = OrderedDict([
    ("DSV4-Pro",                 build_v4_pro_config),
    ("DSV4-Flash",               build_v4_flash_config),
    ("DSV4-Pro-DFlash",          build_dsv4_pro_dflash_config),
    ("DSV4-Pro-DFlash-hybrid",   build_dsv4_pro_dflash_hybrid_config),
    ("DSV4-Flash-DFlash",        build_dsv4_flash_dflash_config),
    ("DSV4-Flash-DFlash-hybrid", build_dsv4_flash_dflash_hybrid_config),
])

DEFAULT_BATCHES = [1, 8, 64, 256]
DEFAULT_SEQ_LENS = [8192, 131072, 262144, 1048576, 4194304]
DEFAULT_HARDWARE = ["H100", "Ascend950PR"]
DEFAULT_DTYPES = {"bf16": torch.bfloat16, "fp16": torch.float16, "fp32": torch.float32}
MODES = ("prefill", "mini-prefill", "decode")


# --------------------------------------------------------------------------- #
# Sweep driver
# --------------------------------------------------------------------------- #

def _build_my_device() -> object:
    """Match the speculative custom hardware used in the example files so
    'my-device' rows in this sweep cross-reference cleanly."""
    return CustomHardware(
        name="my-device",
        hbm_bandwidth=8e12,
        fp16=1000e12,
        bf16=1000e12,
        fp8=2000e12,
    )


def _resolve_hardware(name_or_special: str):
    """Accept preset names + the keyword 'my-device' for the placeholder."""
    if name_or_special == "my-device":
        return _build_my_device()
    return get_hardware(name_or_special)


def _run_one(
    config,
    batch: int,
    seq_len: int,
    mode: str,
    hardware,
    dtype: torch.dtype,
    chunk_len: int,
) -> RooflineReport:
    if mode == "prefill":
        return roofline_prefill(config, batch=batch, seq_len=seq_len, dtype=dtype, hardware=hardware)
    if mode == "mini-prefill":
        return roofline_mini_prefill(
            config, batch=batch, chunk_len=chunk_len, kv_cache_len=seq_len,
            dtype=dtype, hardware=hardware,
        )
    if mode == "decode":
        return roofline_decode(config, batch=batch, kv_cache_len=seq_len, dtype=dtype, hardware=hardware)
    raise ValueError(f"unknown mode: {mode!r}")


def analyze_workloads(
    models: "OrderedDict[str, ModelBuilder]",
    hardware_names: list[str],
    batches: list[int],
    seq_lens: list[int],
    dtype: torch.dtype,
    *,
    modes: tuple[str, ...] = MODES,
    chunk_len: int = 512,
) -> dict[tuple[str, str, int, int, str], RooflineReport]:
    """Run the full sweep. Returns ``{(model, hw, batch, seq_len, mode): report}``.

    Each model builder is invoked once and the resulting config is reused
    across all (hw, batch, seq_len, mode) variants — roofline doesn't
    mutate the config so this is safe and saves construction overhead.
    """
    results: dict[tuple[str, str, int, int, str], RooflineReport] = {}
    cfgs = {name: builder() for name, builder in models.items()}
    for hw_name in hardware_names:
        hw = _resolve_hardware(hw_name)
        for model_name, config in cfgs.items():
            for B in batches:
                for S in seq_lens:
                    for mode in modes:
                        report = _run_one(config, B, S, mode, hw, dtype, chunk_len)
                        results[(model_name, hw_name, B, S, mode)] = report
    return results


# --------------------------------------------------------------------------- #
# Formatting helpers
# --------------------------------------------------------------------------- #

def _scale_time(t: float) -> str:
    if t == 0:
        return "0 s"
    for prefix, threshold in (("s", 1.0), ("ms", 1e-3), ("us", 1e-6), ("ns", 1e-9)):
        if t >= threshold:
            return f"{t / threshold:.2f} {prefix}"
    return f"{t * 1e9:.2f} ns"


def _scale_count(n: float, unit: str = "") -> str:
    """Compact scaled number (e.g. ``12.3 G`` / ``4.5 M``)."""
    if n == 0:
        return f"0 {unit}".rstrip()
    for prefix, threshold in (("T", 1e12), ("G", 1e9), ("M", 1e6), ("K", 1e3)):
        if abs(n) >= threshold:
            return f"{n / threshold:.2f} {prefix}{unit}".rstrip()
    return f"{n:.2f} {unit}".rstrip()


def _scale_seq(s: int) -> str:
    """Compact seq-len label: 8192 -> '8K', 131072 -> '128K', 4194304 -> '4M'."""
    if s % (1024 * 1024) == 0:
        return f"{s // (1024 * 1024)}M"
    if s % 1024 == 0:
        return f"{s // 1024}K"
    return str(s)


def _bound_letter(report: RooflineReport) -> str:
    return "C" if report.bottleneck == "compute" else "M"


def _ridge_point(hw, dtype: torch.dtype) -> float:
    peak = hw.peak_flops.get(dtype, 0.0)
    return peak / hw.hbm_bandwidth if hw.hbm_bandwidth > 0 else 0.0


def _hardware_header(hw_names: list[str], dtype: torch.dtype) -> None:
    print("=" * 78)
    print(f"Hardware lineup (dtype = {dtype})")
    print("-" * 78)
    print(f"{'hardware':18s} {'peak (TFLOPS)':>14s} {'HBM (TB/s)':>12s} {'ridge (F/B)':>14s}")
    for name in hw_names:
        hw = _resolve_hardware(name)
        peak = hw.peak_flops.get(dtype, 0.0)
        ridge = _ridge_point(hw, dtype)
        print(f"  {hw.name:16s} {peak/1e12:14.2f} {hw.hbm_bandwidth/1e12:12.3f} {ridge:14.1f}")
    print("=" * 78)
    print()


def _summary_table(
    results: dict,
    hw_name: str,
    mode: str,
    models: list[str],
    batches: list[int],
    seq_lens: list[int],
) -> None:
    """One (hardware, mode) block: rows = (model, batch), cols = seq_len.

    Cell content is ``time / tok-per-s / bound``.
    """
    print(f"--- {hw_name} | {mode} (cell = time / tok-per-s / bound) ---")
    col_seq = [_scale_seq(s) for s in seq_lens]
    header = f"  {'model':28s} {'B':>4s}  " + "  ".join(f"{c:>20s}" for c in col_seq)
    print(header)
    print("-" * len(header))
    for model in models:
        for b in batches:
            cells = []
            for s in seq_lens:
                r = results[(model, hw_name, b, s, mode)]
                bound = _bound_letter(r)
                t = _scale_time(r.roofline_time_s)
                tps = _scale_count(r.tokens_per_sec, "tok/s")
                cell = f"{t}/{tps}/{bound}"
                cells.append(f"{cell:>20s}")
            print(f"  {model:28s} {b:>4d}  " + "  ".join(cells))
        print()


def _component_block(
    results: dict,
    hw_name: str,
    model: str,
    batch: int,
    seq_len: int,
    mode: str,
) -> None:
    """One (hardware, model, batch, seq_len, mode) per-component table."""
    r = results[(model, hw_name, batch, seq_len, mode)]
    print(f"\n  [{model} | {hw_name} | B={batch} | S={_scale_seq(seq_len)} | {mode}]")
    print(f"  {'component':36s} {'FLOPs':>10s} {'bytes':>10s} {'AI':>8s} {'bound':>5s} {'time':>10s}")
    print(f"  {'-' * 36} {'-' * 10} {'-' * 10} {'-' * 8} {'-' * 5} {'-' * 10}")
    for m in r.modules:
        bound = r._module_bound(m)
        t = r._module_time_s(m)
        ai = m.flops / m.bytes if m.bytes > 0 else 0.0
        kind_name = f"{m.kind}:{m.name}" if m.kind in ("mixer", "ffn", "mhc") else m.name
        print(
            f"  {kind_name:36s} "
            f"{_scale_count(m.flops):>10s} "
            f"{_scale_count(m.bytes):>10s} "
            f"{ai:>8.1f} "
            f"{bound:>5s} "
            f"{_scale_time(t):>10s}"
        )
    print(
        f"  {'TOTAL':36s} "
        f"{_scale_count(r.total_flops):>10s} "
        f"{_scale_count(r.total_bytes):>10s} "
        f"{r.arithmetic_intensity:>8.1f} "
        f"{_bound_letter(r):>5s} "
        f"{_scale_time(r.roofline_time_s):>10s}   {_scale_count(r.tokens_per_sec, 'tok/s')}"
    )


def _dump_xlsx(
    results: dict,
    path: str,
    hardware_names: list[str],
    models: list[str],
    batches: list[int],
    seq_lens: list[int],
    dtype: torch.dtype,
) -> None:
    """Write an Excel workbook readable in Excel / LibreOffice / Google Sheets.

    Layout:
      * Sheet ``README`` — sweep parameters, hardware specs (peak, HBM,
        ridge point), color-coding legend.
      * One sheet per ``(hardware, mode)`` named ``{hw}_{mode}``. Rows
        are ``(model, batch)``; columns are seq_lens; each cell holds
        the wall-clock time + tok/s on two lines, with cell background
        coloured by bottleneck (M = light red, C = light green).
      * Sheet ``totals`` — flat per-(model, hw, B, S, mode) table for
        downstream pivot / filter work.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    MEM_FILL = PatternFill("solid", fgColor="FCE4E4")     # light red
    COMP_FILL = PatternFill("solid", fgColor="E4F4E4")    # light green
    HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")  # light blue
    HEADER_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    # default sheet → README
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "DSV4 + DFlash roofline workload sweep"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"dtype: {dtype}"
    ws["A3"] = f"models  : {', '.join(models)}"
    ws["A4"] = f"batches : {batches}"
    ws["A5"] = f"seq_lens: {seq_lens}  (labels: {[_scale_seq(s) for s in seq_lens]})"

    ws["A7"] = "Hardware"
    ws["B7"] = "Peak (TFLOPS)"
    ws["C7"] = "HBM (TB/s)"
    ws["D7"] = "Ridge point (FLOPs / byte)"
    for col in "ABCD":
        ws[f"{col}7"].font = HEADER_FONT
        ws[f"{col}7"].fill = HEADER_FILL
        ws[f"{col}7"].alignment = CENTER
    for i, hw_name in enumerate(hardware_names):
        hw = _resolve_hardware(hw_name)
        peak = hw.peak_flops.get(dtype, 0.0)
        ridge = _ridge_point(hw, dtype)
        row = 8 + i
        ws.cell(row=row, column=1, value=hw.name)
        ws.cell(row=row, column=2, value=round(peak / 1e12, 2))
        ws.cell(row=row, column=3, value=round(hw.hbm_bandwidth / 1e12, 3))
        ws.cell(row=row, column=4, value=round(ridge, 1))

    legend_start = 10 + len(hardware_names)
    ws.cell(row=legend_start, column=1, value="Cell colour legend")
    ws.cell(row=legend_start, column=1).font = Font(bold=True)
    ws.cell(row=legend_start + 1, column=1, value="memory-bound").fill = MEM_FILL
    ws.cell(row=legend_start + 2, column=1, value="compute-bound").fill = COMP_FILL
    ws.cell(row=legend_start + 4, column=1, value=(
        "Cell content: ``time`` on top line, ``tok/s`` below. Bound classification "
        "comes from max(compute_time, memory_time) — whichever wins is the cell's bound."
    ))
    ws.cell(row=legend_start + 4, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=legend_start + 4, start_column=1, end_row=legend_start + 4, end_column=7)

    for col_letter, width in zip("ABCDE", (28, 18, 14, 28, 18)):
        ws.column_dimensions[col_letter].width = width

    # --- One sheet per (hw, mode) ---
    for hw_name in hardware_names:
        for mode in MODES:
            sheet_name = f"{hw_name}_{mode}"[:31]  # Excel max sheet name 31 chars
            ws = wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=f"{hw_name} | {mode}").font = Font(bold=True, size=12)

            # Header row at row 3
            header_row = 3
            ws.cell(row=header_row, column=1, value="model")
            ws.cell(row=header_row, column=2, value="batch")
            for j, s in enumerate(seq_lens):
                ws.cell(row=header_row, column=3 + j, value=_scale_seq(s))
            for col in range(1, 3 + len(seq_lens)):
                c = ws.cell(row=header_row, column=col)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = CENTER
                c.border = BORDER

            # Data rows
            row_idx = header_row + 1
            for model in models:
                for B in batches:
                    ws.cell(row=row_idx, column=1, value=model).alignment = LEFT
                    ws.cell(row=row_idx, column=2, value=B).alignment = CENTER
                    for j, s in enumerate(seq_lens):
                        r = results[(model, hw_name, B, s, mode)]
                        bound = _bound_letter(r)
                        fill = MEM_FILL if bound == "M" else COMP_FILL
                        # Two lines: time + tok/s. Wrap_text makes the newline render.
                        text = f"{_scale_time(r.roofline_time_s)}\n{_scale_count(r.tokens_per_sec, 'tok/s')}"
                        c = ws.cell(row=row_idx, column=3 + j, value=text)
                        c.fill = fill
                        c.alignment = CENTER
                        c.border = BORDER
                    row_idx += 1
                # Visual gap between models — empty row with thin separator look.
                row_idx += 1

            # Freeze the header rows + first two cols so scrolling stays legible.
            ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

            # Column widths: model + batch + seq cols.
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 7
            for j in range(len(seq_lens)):
                ws.column_dimensions[get_column_letter(3 + j)].width = 16
            # Row heights big enough for 2-line cells.
            for r_i in range(header_row + 1, row_idx):
                ws.row_dimensions[r_i].height = 28

    # --- Totals sheet (flat) for pivot tables / filters ---
    ws = wb.create_sheet("totals")
    headers = ["model", "hardware", "batch", "seq_len", "mode",
               "total_flops", "total_bytes", "AI", "compute_time_s", "memory_time_s",
               "roofline_time_s", "tokens_per_sec", "bottleneck"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    r_i = 2
    for (model, hw, B, S, mode), r in results.items():
        ws.cell(row=r_i, column=1, value=model)
        ws.cell(row=r_i, column=2, value=hw)
        ws.cell(row=r_i, column=3, value=B)
        ws.cell(row=r_i, column=4, value=S)
        ws.cell(row=r_i, column=5, value=mode)
        ws.cell(row=r_i, column=6, value=r.total_flops)
        ws.cell(row=r_i, column=7, value=r.total_bytes)
        ws.cell(row=r_i, column=8, value=round(r.arithmetic_intensity, 3))
        ws.cell(row=r_i, column=9, value=r.compute_time_s)
        ws.cell(row=r_i, column=10, value=r.memory_time_s)
        ws.cell(row=r_i, column=11, value=r.roofline_time_s)
        ws.cell(row=r_i, column=12, value=round(r.tokens_per_sec, 2))
        ws.cell(row=r_i, column=13, value=r.bottleneck)
        r_i += 1
    ws.auto_filter.ref = f"A1:M{r_i - 1}"  # turn on filter
    ws.freeze_panes = "F2"
    for j, width in enumerate((28, 14, 7, 12, 14, 16, 16, 8, 14, 14, 14, 14, 12), 1):
        ws.column_dimensions[get_column_letter(j)].width = width

    wb.save(path)
    print(f"\nExcel written: {path}")


def _dump_csv(results: dict, path: str) -> None:
    """Write two CSV files: model-totals + per-component."""
    totals_path = Path(path)
    comps_path = totals_path.with_suffix(totals_path.suffix + ".components.csv") if totals_path.suffix \
        else totals_path.with_name(totals_path.name + ".components.csv")

    # Totals: one row per (model, hw, B, S, mode)
    with totals_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "model", "hardware", "batch", "seq_len", "mode",
            "total_flops", "total_bytes", "AI", "compute_time_s", "memory_time_s",
            "roofline_time_s", "tokens_per_sec", "bottleneck",
        ])
        for (model, hw, B, S, mode), r in results.items():
            w.writerow([
                model, hw, B, S, mode,
                r.total_flops, r.total_bytes, f"{r.arithmetic_intensity:.4f}",
                f"{r.compute_time_s:.6e}", f"{r.memory_time_s:.6e}",
                f"{r.roofline_time_s:.6e}", f"{r.tokens_per_sec:.4f}",
                r.bottleneck,
            ])

    # Components: one row per ModuleStat
    with comps_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "model", "hardware", "batch", "seq_len", "mode",
            "kind", "name", "layer_idx", "flops", "bytes", "AI", "bound", "time_s",
        ])
        for (model, hw, B, S, mode), r in results.items():
            for m in r.modules:
                ai = m.flops / m.bytes if m.bytes > 0 else 0.0
                w.writerow([
                    model, hw, B, S, mode,
                    m.kind, m.name,
                    "" if m.layer_idx is None else m.layer_idx,
                    m.flops, m.bytes, f"{ai:.4f}",
                    r._module_bound(m),
                    f"{r._module_time_s(m):.6e}",
                ])

    print(f"\nCSV written: {totals_path}")
    print(f"CSV written: {comps_path}")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def _parse_csv_list(s: str) -> list[str]:
    return [x.strip() for x in s.split(",") if x.strip()]


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--models",
        default=",".join(ALL_MODELS.keys()),
        help=f"Comma-separated subset of {list(ALL_MODELS.keys())}.",
    )
    parser.add_argument(
        "--hardware", "--hw",
        default=",".join(DEFAULT_HARDWARE + ["my-device"]),
        help="Comma-separated preset names (or 'my-device' for the example placeholder).",
    )
    parser.add_argument(
        "--batches",
        default=",".join(str(b) for b in DEFAULT_BATCHES),
        help="Comma-separated batch sizes.",
    )
    parser.add_argument(
        "--seq-lens",
        default=",".join(str(s) for s in DEFAULT_SEQ_LENS),
        help="Comma-separated sequence lengths (used as Q for prefill, kv_cache for the others).",
    )
    parser.add_argument(
        "--chunk-len", type=int, default=512,
        help="Chunked-prefill chunk length Q (mini-prefill mode). Default 512.",
    )
    parser.add_argument(
        "--dtype", default="bf16", choices=list(DEFAULT_DTYPES.keys()),
    )
    parser.add_argument(
        "--components", action="store_true",
        help="Also print per-component breakdowns (one block per "
             "(hw, model, B, S, mode); can be a lot of output — narrow the sweep first).",
    )
    parser.add_argument(
        "--csv", default=None,
        help="Path prefix to dump full CSV (totals + components).",
    )
    parser.add_argument(
        "--xlsx", default=None,
        help="Path to dump an Excel workbook (one sheet per (hardware, mode) "
             "plus a README sheet and a flat 'totals' sheet for filtering).",
    )
    args = parser.parse_args()

    models_subset = OrderedDict(
        (m, ALL_MODELS[m]) for m in _parse_csv_list(args.models) if m in ALL_MODELS
    )
    if not models_subset:
        print(f"No valid models in --models={args.models}; choose from {list(ALL_MODELS)}")
        return 1
    hardware_names = _parse_csv_list(args.hardware)
    batches = [int(x) for x in _parse_csv_list(args.batches)]
    seq_lens = [int(x) for x in _parse_csv_list(args.seq_lens)]
    dtype = DEFAULT_DTYPES[args.dtype]

    print(f"models   : {list(models_subset.keys())}")
    print(f"hardware : {hardware_names}")
    print(f"batches  : {batches}")
    print(f"seq_lens : {seq_lens}  (labels: {[_scale_seq(s) for s in seq_lens]})")
    print(f"chunk_len: {args.chunk_len}  (mini-prefill Q)")
    print(f"dtype    : {dtype}")
    print()

    results = analyze_workloads(
        models_subset, hardware_names, batches, seq_lens, dtype,
        chunk_len=args.chunk_len,
    )

    _hardware_header(hardware_names, dtype)

    # --- Per-(hardware, mode) summary matrices ---
    for hw_name in hardware_names:
        print(f"================ {hw_name} ================")
        for mode in MODES:
            _summary_table(
                results, hw_name, mode,
                list(models_subset.keys()), batches, seq_lens,
            )
        print()

    # --- Optional per-component dumps ---
    if args.components:
        print("=" * 78)
        print("PER-COMPONENT BREAKDOWN")
        print("=" * 78)
        for hw_name in hardware_names:
            for model in models_subset.keys():
                for b in batches:
                    for s in seq_lens:
                        for mode in MODES:
                            _component_block(results, hw_name, model, b, s, mode)
        print()

    if args.csv:
        _dump_csv(results, args.csv)

    if args.xlsx:
        _dump_xlsx(
            results, args.xlsx, hardware_names,
            list(models_subset.keys()), batches, seq_lens, dtype,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
