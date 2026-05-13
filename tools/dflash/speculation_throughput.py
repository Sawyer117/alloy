"""DFlash speculation throughput sweep — (main, draft) pairs × workload axes.

Companion to ``analyze_workloads.py``. Where ``analyze_workloads`` sweeps
per-model prefill / decode, this script sweeps the **speculation-round**
cost (draft forward + target verify) for fixed main + draft pairs across
batches, past-cache lengths, and acceptance rates, and reports effective
tok/s alongside the autoregressive (target-only) baseline.

Outputs (in order of usefulness):
  * Per-(pair, hardware) Excel sheet with effective tok/s + speedup vs AR
    matrix. Rows = (batch, avg_accept_len), cols = past_cache_len.
  * Flat 'rounds' sheet with full per-row data for filter / pivot.
  * Auto-coloured cells: green = speedup >= 1 (DFlash wins), red < 1.

Usage::

    python -m alloy.tools.dflash.speculation_throughput \\
        --xlsx D:/work/model_gym/dflash_speculation.xlsx

    # narrow scope:
    python -m alloy.tools.dflash.speculation_throughput \\
        --pairs DSV4-Flash:DSV4-Flash-DFlash-hybrid \\
        --batches 1 --past-cache 8192,131072 --accepts 1.0,2.0 \\
        --hardware H100
"""
from __future__ import annotations

import argparse
import sys
from collections import OrderedDict
from pathlib import Path

import torch

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from alloy.roofline import (
    CustomHardware,
    get_hardware,
    roofline_dflash_steady_throughput,
)

from alloy.examples.roofline.dsv4_pro import build_v4_pro_config
from alloy.examples.roofline.dsv4_flash import build_v4_flash_config
from alloy.examples.roofline.dflash.dsv4_pro_dflash import (
    build_dsv4_pro_dflash_config,
)
from alloy.examples.roofline.dflash.dsv4_pro_dflash_hybrid import (
    build_dsv4_pro_dflash_hybrid_config,
)
from alloy.examples.roofline.dflash.dsv4_flash_dflash import (
    build_dsv4_flash_dflash_config,
)
from alloy.examples.roofline.dflash.dsv4_flash_dflash_hybrid import (
    build_dsv4_flash_dflash_hybrid_config,
)


# --------------------------------------------------------------------------- #
# Defaults
# --------------------------------------------------------------------------- #

# Available main-draft pairings. Naming convention is ``MAIN:DRAFT``.
ALL_PAIRS: "OrderedDict[str, tuple]" = OrderedDict([
    ("DSV4-Pro:DSV4-Pro-DFlash",
        (build_v4_pro_config, build_dsv4_pro_dflash_config)),
    ("DSV4-Pro:DSV4-Pro-DFlash-hybrid",
        (build_v4_pro_config, build_dsv4_pro_dflash_hybrid_config)),
    ("DSV4-Flash:DSV4-Flash-DFlash",
        (build_v4_flash_config, build_dsv4_flash_dflash_config)),
    ("DSV4-Flash:DSV4-Flash-DFlash-hybrid",
        (build_v4_flash_config, build_dsv4_flash_dflash_hybrid_config)),
])

DEFAULT_BATCHES = [1, 8, 64]
DEFAULT_PAST_CACHE = [8192, 32768, 131072, 1048576]
DEFAULT_ACCEPTS = [0.7, 1.44, 2.44]   # 0.7 = MTP=1, 1.44 = MTP=3 from SKILL.md
DEFAULT_HARDWARE = ["H100", "Ascend950PR"]


def _resolve_hardware(name: str):
    if name == "my-device":
        return CustomHardware(
            name="my-device", hbm_bandwidth=8e12,
            fp16=1000e12, bf16=1000e12, fp8=2000e12,
        )
    return get_hardware(name)


def _parse_csv_list(s: str) -> list[str]:
    return [x.strip() for x in s.split(",") if x.strip()]


def sweep(
    pairs: "OrderedDict[str, tuple]",
    hardware_names: list[str],
    batches: list[int],
    past_caches: list[int],
    accepts: list[float],
    *,
    block_size: int,
    dtype: torch.dtype,
) -> list[dict]:
    """Run the full grid and return a flat list of result rows."""
    rows = []
    cfgs = {name: (mb(), db()) for name, (mb, db) in pairs.items()}
    for hw_name in hardware_names:
        hw = _resolve_hardware(hw_name)
        for pair_name, (target, draft) in cfgs.items():
            for B in batches:
                for P in past_caches:
                    for acc in accepts:
                        s = roofline_dflash_steady_throughput(
                            target, draft,
                            batch=B, past_cache_len=P,
                            block_size=block_size, avg_accept_len=acc,
                            target_hardware=hw, draft_hardware=hw,
                            dtype=dtype,
                        )
                        rows.append({
                            "pair": pair_name,
                            "hardware": hw_name,
                            "batch": B,
                            "past_cache": P,
                            "avg_accept": acc,
                            "block_size": block_size,
                            "t_draft_ms": s["draft"].roofline_time_s * 1000,
                            "t_target_verify_ms": s["target"].roofline_time_s * 1000,
                            "t_round_ms": s["t_round"] * 1000,
                            "t_ar_per_token_ms": s["ar_baseline_t_round"] * 1000,
                            "DFlash_tps": s["effective_tps"],
                            "AR_tps": s["ar_baseline_tps"],
                            "speedup": s["speedup_vs_ar"],
                            "draft_bound": s["draft"].bottleneck,
                            "target_bound": s["target"].bottleneck,
                        })
    return rows


# --------------------------------------------------------------------------- #
# Excel output
# --------------------------------------------------------------------------- #

def _scale_seq(s: int) -> str:
    if s % (1024 * 1024) == 0:
        return f"{s // (1024 * 1024)}M"
    if s % 1024 == 0:
        return f"{s // 1024}K"
    return str(s)


def write_xlsx(
    rows: list[dict],
    path: str,
    pairs: list[str],
    hardware_names: list[str],
    batches: list[int],
    past_caches: list[int],
    accepts: list[float],
    *,
    block_size: int,
    dtype: torch.dtype,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    WIN_FILL = PatternFill("solid", fgColor="D4F4D4")     # speedup >= 1.5
    EQUAL_FILL = PatternFill("solid", fgColor="FFF7CC")   # 1.0 <= speedup < 1.5
    LOSE_FILL = PatternFill("solid", fgColor="FCE4E4")    # < 1.0
    HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
    HEADER_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def speedup_fill(speedup: float):
        if speedup >= 1.5:
            return WIN_FILL
        if speedup >= 1.0:
            return EQUAL_FILL
        return LOSE_FILL

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "DFlash speculation throughput sweep"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"dtype : {dtype}"
    ws["A3"] = f"block_size : {block_size}  (number of speculative tokens per round)"
    ws["A4"] = f"pairs    : {pairs}"
    ws["A5"] = f"batches  : {batches}"
    ws["A6"] = f"past_cache : {past_caches}  (labels: {[_scale_seq(p) for p in past_caches]})"
    ws["A7"] = f"avg_accept : {accepts}"
    ws["A9"] = "Speed metric: effective_tps = B * (avg_accept + 1) / (t_draft + t_target_verify)"
    ws["A10"] = "AR baseline : single-token target decode at the same past_cache_len"
    ws["A11"] = "Cell colours: green = DFlash >= 1.5x AR;  yellow = 1.0x-1.5x;  red = < 1.0x"
    ws["A13"] = "NOTE: alloy roofline uses worst-case 'any-routing' upper bound for MoE expert loading."
    ws["A14"] = ("On 384-expert V4-Pro with Q=16 verify this loads 96 experts and dramatically "
                 "overstates t_target_verify vs. real routing locality. Treat MoE-target speedups "
                 "as a pessimistic floor; real-world routing locality typically improves them.")
    for r in (13, 14):
        ws[f"A{r}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    ws.column_dimensions["A"].width = 100

    # Per-(pair, hardware) sheets — speedup matrix
    for pair_name in pairs:
        for hw_name in hardware_names:
            sheet_name = f"{pair_name.split(':')[0][:6]}_{pair_name.split(':')[1][:10]}_{hw_name[:6]}"[:31]
            ws = wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=f"{pair_name} | {hw_name}").font = Font(bold=True, size=12)
            ws.cell(row=2, column=1, value=(
                "cell: effective tok/s, AR tok/s, speedup — colour by speedup"
            ))

            # Header
            header_row = 4
            ws.cell(row=header_row, column=1, value="batch")
            ws.cell(row=header_row, column=2, value="avg_accept")
            for j, p in enumerate(past_caches):
                ws.cell(row=header_row, column=3 + j, value=f"P={_scale_seq(p)}")
            for col in range(1, 3 + len(past_caches)):
                c = ws.cell(row=header_row, column=col)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = CENTER
                c.border = BORDER

            # Rows
            row_idx = header_row + 1
            for B in batches:
                for acc in accepts:
                    ws.cell(row=row_idx, column=1, value=B).alignment = CENTER
                    ws.cell(row=row_idx, column=2, value=acc).alignment = CENTER
                    for j, P in enumerate(past_caches):
                        row = next(r for r in rows
                                   if r["pair"] == pair_name and r["hardware"] == hw_name
                                   and r["batch"] == B and r["past_cache"] == P
                                   and r["avg_accept"] == acc)
                        text = (
                            f"{row['DFlash_tps']:.0f} tok/s\n"
                            f"AR {row['AR_tps']:.0f} tok/s\n"
                            f"{row['speedup']:.2f}x"
                        )
                        c = ws.cell(row=row_idx, column=3 + j, value=text)
                        c.fill = speedup_fill(row["speedup"])
                        c.alignment = CENTER
                        c.border = BORDER
                    row_idx += 1
                row_idx += 1  # blank between batches

            ws.freeze_panes = ws.cell(row=header_row + 1, column=3)
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 12
            for j in range(len(past_caches)):
                ws.column_dimensions[get_column_letter(3 + j)].width = 18
            for r_i in range(header_row + 1, row_idx):
                ws.row_dimensions[r_i].height = 42

    # Flat 'rounds' sheet
    ws = wb.create_sheet("rounds")
    headers = list(rows[0].keys()) if rows else []
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    for r_i, row in enumerate(rows, 2):
        for j, h in enumerate(headers, 1):
            val = row[h]
            if isinstance(val, float):
                ws.cell(row=r_i, column=j, value=round(val, 4))
            else:
                ws.cell(row=r_i, column=j, value=val)
        # Colour speedup cell
        if "speedup" in headers:
            si = headers.index("speedup") + 1
            ws.cell(row=r_i, column=si).fill = speedup_fill(row["speedup"])
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "F2"
    for j, h in enumerate(headers, 1):
        width = max(12, min(28, len(h) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width

    wb.save(path)
    print(f"Excel written: {path}")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--pairs",
        default=",".join(ALL_PAIRS.keys()),
        help=f"Comma-separated main:draft pairs from {list(ALL_PAIRS.keys())}.",
    )
    parser.add_argument(
        "--hardware", default=",".join(DEFAULT_HARDWARE),
        help="Comma-separated hardware preset names.",
    )
    parser.add_argument(
        "--batches", default=",".join(str(b) for b in DEFAULT_BATCHES),
    )
    parser.add_argument(
        "--past-cache", default=",".join(str(p) for p in DEFAULT_PAST_CACHE),
        help="Past-cache lengths (= confirmed prefix at the round being modeled).",
    )
    parser.add_argument(
        "--accepts", default=",".join(str(a) for a in DEFAULT_ACCEPTS),
        help="Mean accepted tokens per round. SKILL.md table: MTP=1 → 0.7, MTP=3 → 1.44.",
    )
    parser.add_argument("--block-size", type=int, default=16)
    parser.add_argument("--dtype", default="bf16", choices=["bf16", "fp16", "fp32"])
    parser.add_argument(
        "--xlsx", default=None,
        help="Path to write an Excel workbook (README + per-pair-hw sheets + flat rounds sheet).",
    )
    args = parser.parse_args()

    pairs_subset = OrderedDict(
        (p, ALL_PAIRS[p]) for p in _parse_csv_list(args.pairs) if p in ALL_PAIRS
    )
    hardware_names = _parse_csv_list(args.hardware)
    batches = [int(x) for x in _parse_csv_list(args.batches)]
    past_caches = [int(x) for x in _parse_csv_list(args.past_cache)]
    accepts = [float(x) for x in _parse_csv_list(args.accepts)]
    dtype = {"bf16": torch.bfloat16, "fp16": torch.float16, "fp32": torch.float32}[args.dtype]

    print(f"pairs    : {list(pairs_subset.keys())}")
    print(f"hardware : {hardware_names}")
    print(f"batches  : {batches}")
    print(f"past_cache: {past_caches}")
    print(f"accepts  : {accepts}")
    print(f"block_size: {args.block_size}")
    print(f"dtype    : {dtype}")
    print()

    rows = sweep(
        pairs_subset, hardware_names, batches, past_caches, accepts,
        block_size=args.block_size, dtype=dtype,
    )

    # Always print a short summary to stdout
    print(f"{'pair':38s} {'hw':12s} {'B':>3s} {'P':>8s} {'acc':>5s} "
          f"{'t_round':>9s} {'DFlash':>9s} {'AR':>9s} {'speedup':>8s}")
    print("-" * 110)
    for row in rows:
        print(f"{row['pair']:38s} {row['hardware']:12s} "
              f"{row['batch']:>3d} {_scale_seq(row['past_cache']):>8s} "
              f"{row['avg_accept']:>5.2f} {row['t_round_ms']:>7.2f}ms "
              f"{row['DFlash_tps']:>7.1f}/s {row['AR_tps']:>7.1f}/s "
              f"{row['speedup']:>6.2f}x")

    if args.xlsx:
        write_xlsx(
            rows, args.xlsx,
            list(pairs_subset.keys()), hardware_names, batches, past_caches, accepts,
            block_size=args.block_size, dtype=dtype,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
